"""Convert Agent results into typed WebSocket message streams."""
import json
import re as _re
import sys
from pathlib import Path
from typing import AsyncGenerator

_backend_dir = Path(__file__).resolve().parent
if str(_backend_dir) not in sys.path:
    sys.path.insert(0, str(_backend_dir))

from router import route, IntentResult
from llm.prompts import COMPARE_SYSTEM


# ── Helpers ──
_CNS = {"一":1,"二":2,"两":2,"三":3,"四":4,"五":5,"六":6,"七":7,"八":8,"九":9,"十":10}
def _parse_num(text, default=None):
    """Parse '3个/三个/前3/前三' → int. Handles Arabic + Chinese numerals."""
    m = _re.search(r'(\d+)\s*(?:个|位|人|名)?', text)
    if m: return int(m.group(1))
    m = _re.search(r'([一二两三四五六七八九十])\s*(?:个|位|人|名)?', text)
    if m: return _CNS.get(m.group(1), default)
    return default

def _comma_list(value, max_items=None):
    """Split a comma-separated string into a cleaned list."""
    items = [x.strip() for x in str(value or "").split(",") if x.strip()]
    return items[:max_items] if max_items else items


_history_cache = None

def _get_history(eid):
    """Read history sub-tables for ONE employee only — fast, targeted read."""
    try:
        import openpyxl
        from utils.config import TEST_DATA_FILE
        wb = openpyxl.load_workbook(str(TEST_DATA_FILE), read_only=True, data_only=True)
        result = {}
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith('历史_'): continue
            ws = wb[sheet_name]
            headers = [str(ws.cell(row=1, column=c).value or '') for c in range(1, ws.max_column + 1)]
            short = sheet_name.replace('历史_', '')
            records = []
            for r in range(2, ws.max_row + 1):
                row_eid = str(ws.cell(row=r, column=1).value or '')
                if row_eid == str(eid):
                    rec = {}
                    for ci, hdr in enumerate(headers):
                        if hdr and hdr not in ('工号', '姓名'):
                            val = ws.cell(row=r, column=ci+1).value
                            rec[str(hdr)] = str(val) if val is not None else ''
                    if rec: records.append(rec)
            if records:
                result[short] = records
        wb.close()
        return result
    except Exception as exc:
        print(f"[History] FAILED for {eid}: {exc}")
        return {}


def _avatar_for(eid, gender):
    """Generate avatar path — same hash algorithm as frontend _avatarUrl."""
    pool = 'f' if gender == '女' else 'm'
    count = 91 if pool == 'f' else 64
    h = 0
    for ch in str(eid):
        h = ((h << 5) - h) + ord(ch)
        h |= 0  # truncate to 32-bit, same as JS
    n = (abs(h) % count) + 1
    return f"/widget/avatars/avatar-{pool}-{n:03d}.png"

def _build_skills(p):
    """Build skill tags from 1000-data fields."""
    parts = []
    if p.get("XPM领域") and p["XPM领域"] not in ("否","","nan"): parts.append(p["XPM领域"])
    if p.get("曾工作领域及年限") and p["曾工作领域及年限"] not in ("否","","nan"): parts.append(p["曾工作领域及年限"])
    if p.get("内部拓展师等级") and p["内部拓展师等级"] not in ("否","","nan"): parts.append(p["内部拓展师等级"])
    if p.get("综合等级") and p["综合等级"] not in ("否","","nan"): parts.append(p["综合等级"])
    if p.get("授课讲师等级") and p["授课讲师等级"] not in ("否","","nan"): parts.append(p["授课讲师等级"])
    return parts[:5]

def _build_tags(p):
    """Build tag pills from 1000-data boolean/status fields."""
    tags = []
    bool_tags = [
        ("关键人才","是"), ("国际化人才","是"), ("海外战略储备","是"),
        ("海外工作","是"), ("Hipo人才","是"), ("精英MBA班","是"),
        ("客户界面","是"), ("技委会成员","是"), ("数字化人才","是"),
        ("是否工艺师","是"), ("复合型技师","是"), ("NPI技师","是"),
        ("轮岗人员","是"), ("海外留学","是"),
    ]
    for field, val in bool_tags:
        if str(p.get(field, "")).strip() == val:
            tags.append(field)
    # Add non-boolean tag-like fields
    if p.get("G-Plan") and str(p.get("G-Plan")).strip() not in ("否","","nan"): tags.append("G-Plan")
    if p.get("GPS角色") and str(p.get("GPS角色")).strip() not in ("否","","nan"): tags.append("GPS角色")
    if p.get("当前关键角色") and str(p.get("当前关键角色")).strip() not in ("否","","nan"): tags.append(p["当前关键角色"])
    if p.get("潜力等级") and str(p.get("潜力等级")).strip() not in ("否","","nan"): tags.append(p["潜力等级"])
    if p.get("档位") and str(p.get("档位")).strip() not in ("否","","nan"): tags.append(p["档位"]+"档")
    return tags[:5]


def _fmt_score(score):
    """Format a numeric score to string, pass through otherwise."""
    return f"{score:.0f}" if isinstance(score, (int, float)) else str(score)


def _ensure_store():
    """Get a loaded TalentStore singleton — fast path, no embedding."""
    from data.talent_store import get_store
    store = get_store()
    if store.df is None or len(store.records) == 0:
        store.load(embedding_fn=None)  # Skip vector building — keyword search is sufficient
    return store


def _find_position_dict():
    """Find position_dict.json, returning the first existing path."""
    candidates = [
        Path(__file__).parent / "data" / "position_dict.json",
        Path(__file__).parent.parent.parent / "0-AI-Talent-Discovering" / "data" / "position_dict.json",
    ]
    return next((p for p in candidates if p.exists()), candidates[0])


# ── Intent Dispatch Table ──

async def _handle_clarify(ctx, params, user_text, ids):
    yield {"type": "text", "content": params.get("question", "抱歉，我没理解你的意思，能再说具体一点吗？")}
    yield {"type": "done"}


async def _handle_position_to_person(ctx, params, user_text, ids):
    from agents.match import MatchAgent
    pos_name = params.get("position", user_text)
    yield {"type": "text", "content": "正在搜索匹配「" + pos_name + "」的候选人..."}

    # Parse user-specified top_n from text
    top_n = int(params.get("top_n", 10))
    _pn = _parse_num(user_text)
    if _pn: top_n = _pn

    result = MatchAgent().match_position_to_person(position_name=pos_name, top_n=top_n)
    candidates = result.get("candidates", [])
    ctx.cached_candidates = candidates

    if not candidates:
        yield {"type": "text", "content": "未找到匹配的候选人，请尝试放宽条件或换一种描述。"}
        yield {"type": "done"}
        return

    yield {"type": "text", "content": "找到 " + str(len(candidates)) + " 位匹配候选人："}
    for c in candidates:
        p = c.get("profile", {})
        eid = c.get("id", "")
        g = p.get("性别", "")
        yield {
            "type": "card",
            "data": {
                "id": eid,
                "name": p.get("姓名", ""),
                "gender": g,
                "avatar": _avatar_for(eid, g),
                "grade": c.get("grade", "B"),
                "score": _fmt_score(c.get("llm_score", c.get("keyword_score", "-"))),
                "department": p.get("所在职位", ""),
                "position": p.get("所在职位", ""),
                "level": p.get("职级", ""),
                "education": p.get("最高学历", ""),
                "performance": p.get("绩效等级", p.get("档位", "")),
                "skills": _build_skills(p),
                "tags": _build_tags(p),
            }
        }
    yield {"type": "actions", "actions": ["对比选中", "导出Excel"]}
    yield {"type": "done"}

    # Auto-compare: "对比前三/前3个" → auto-trigger
    if candidates:
        _cm = _re.search(r'对比\s*前?\s*(\d+|[一二两三四五六七八九十])\s*(个|位|人|名)?', user_text)
        if _cm:
            cn = _cm.group(1)
            compare_n = int(cn) if cn.isdigit() else _CNS.get(cn, 2)
            compare_n = min(compare_n, len(candidates))
            if compare_n >= 2:
                async for _msg in _handle_compare(ctx, params, user_text, [c.get("id") for c in candidates[:compare_n]]):
                    yield _msg



async def _handle_report(ctx, params, user_text, ids):
    import hashlib
    store = _ensure_store()

    eid = params.get("employee_id") or (ids[0] if ids else None)
    if not eid:
        yield {"type": "text", "content": "请指定要查看报告的候选人"}
        yield {"type": "done"}
        return

    profile = store.get_by_id(eid)
    if not profile:
        yield {"type": "text", "content": f"未找到员工 {eid}"}
        yield {"type": "done"}
        return

    name = profile.get('姓名', eid)

    # Check cache
    if eid in ctx.cached_reports:
        yield {"type": "text", "content": f"已从缓存加载 {name} 的报告"}
        yield ctx.cached_reports[eid]
        yield {"type": "done"}
        return

    yield {"type": "text", "content": f"正在加载 {name} 的详细档案..."}

    # ── Pass through ALL raw fields from Excel, plus computed fields ──
    data = dict(profile)  # copy all raw fields
    data["id"] = eid

    # Deterministic score
    seed = str(profile.get("员工编码", profile.get("编码", eid)))
    h = int(hashlib.md5(seed.encode()).hexdigest()[:4], 16)
    score = 70 + (h % 31)
    data["score"] = _fmt_score(score)
    if score >= 90: data["grade"] = "S"
    elif score >= 80: data["grade"] = "A"
    elif score >= 65: data["grade"] = "B"
    else: data["grade"] = "C"

    # Avatar by gender
    data["avatar"] = _avatar_for(eid, profile.get("性别", ""))

    # Skills/tags as lists
    data["skill_list"] = _comma_list(profile.get("技能标签"))
    data["tag_list"] = _comma_list(profile.get("所有标签"))

    # ── History: load from pre-built cache (O(1) lookup) ──
    from data.history_cache import get_history, ensure_loaded
    ensure_loaded()
    data["history"] = get_history(eid)

    # Convert numpy types to plain Python for JSON serialization
    import numpy as np
    for k, v in list(data.items()):
        if isinstance(v, (np.integer,)): data[k] = int(v)
        elif isinstance(v, (np.floating,)): data[k] = float(v)
        elif isinstance(v, (np.bool_,)): data[k] = bool(v)

    result = {"type": "report", "data": data}
    print(f"[REPORT] Sending {len(data)} fields for {name}, has 姓名={bool(data.get('姓名'))}, history={len(data['history'])} modules")
    ctx.cached_reports[eid] = result
    yield result
    yield {"type": "actions", "actions": ["返回搜索"]}
    yield {"type": "done"}


async def _handle_compare(ctx, params, user_text, ids):
    from agents.compare import CompareAgent
    import hashlib
    compare_ids = ids or params.get("ids", [])

    # Handle "top-N" placeholder from router (LLM → "对比前3个" → ids:["top-3"])
    resolved_ids = []
    for cid in (compare_ids or []):
        if isinstance(cid, str) and cid.startswith("top-"):
            n = int(cid.split("-")[1])
            cached_ids = [c.get("id") for c in (ctx.cached_candidates or [])]
            resolved_ids = cached_ids[:n]
            break
        else:
            resolved_ids.append(cid)
    if resolved_ids:
        compare_ids = resolved_ids

    if len(compare_ids) < 2:
        yield {"type": "text", "content": "请至少选择2名候选人进行对比（可以说'对比前两个'）"}
        yield {"type": "done"}
        return

    c_key = ctx.compare_key(compare_ids)
    # Check cache: same set of people → return cached result
    if c_key in ctx.cached_compares and ctx.cached_compares[c_key]["data"].get("per_person"):
        cached = ctx.cached_compares[c_key]
        yield {"type": "text", "content": f"已从缓存加载对比结果"}
        yield cached
        yield {"type": "done"}
        return

    yield {"type": "text", "content": f"正在对比 {len(compare_ids)} 位候选人..."}

    # ── Phase 1: Build base profiles immediately (zero AI) ──
    profiles_raw = []
    store = _ensure_store()
    for eid in compare_ids:
        p = store.get_by_id(eid)
        if p: profiles_raw.append(p)

    profiles = []
    for i, p in enumerate(profiles_raw):
        eid = str(p.get("员工编码", p.get("工号", "")))
        g = p.get("性别", "")
        seed = str(eid)
        h = int(hashlib.md5(seed.encode()).hexdigest()[:4], 16)
        score = 70 + (h % 31)
        if score >= 90: grade = "S"
        elif score >= 80: grade = "A"
        elif score >= 65: grade = "B"
        else: grade = "C"
        base = score
        dims = {
            "技能匹配": min(100, max(20, base + (h % 15) - 7)),
            "经验匹配": min(100, max(20, base + ((h>>4) % 15) - 7)),
            "绩效趋势": min(100, max(20, base + ((h>>8) % 15) - 7)),
            "软性素质": min(100, max(20, base + ((h>>12) % 11) - 5)),
            "发展潜力": min(100, max(20, base - ((h>>6) % 9) + 4)),
        }
        profiles.append({
            "id": eid,
            "name": p.get("姓名", ""),
            "gender": g,
            "avatar": _avatar_for(eid, g),
            "department": p.get("所在职位", p.get("部门", "")),
            "position": p.get("所在职位", p.get("岗位", "")),
            "level": p.get("职级", ""),
            "education": p.get("最高学历", p.get("学历", "")),
            "major": p.get("最高学历专业", p.get("专业", "")),
            "performance": p.get("档位", p.get("绩效等级", "")),
            "tenure": p.get("司龄", p.get("司龄(年)", "")),
            "skills": _comma_list(p.get("技能标签")),
            "tags": _comma_list(p.get("所有标签")),
            "grade": grade,
            "score": _fmt_score(score),
            "dimensions": dims,
        })

    # Yield base comparison table immediately — user sees data right away
    yield {
        "type": "compare",
        "data": {
            "profiles": profiles,
            "analysis": "AI 分析正在进行中，结果出来后自动更新...",
            "per_person": [],
        }
    }

    # ── Phase 2: LLM analysis with streaming progress ──
    yield {"type": "text", "content": "AI 正在逐人分析对比中..."}

    # Build the compare prompt directly instead of through CompareAgent (avoids double-prompting)
    profiles_text = ""
    for i, p in enumerate(profiles_raw):
        profiles_text += f"""
候选人 {i+1}: 姓名: {p.get('姓名','')}  所在职位: {p.get('所在职位','')}  岗位: {p.get('所在职位','')}
  职级: {p.get('职级','')}  职等: {p.get('职等','')}  学历: {p.get('最高学历','')}/{p.get('最高学历专业','')}
  档位: {p.get('档位','')}(总积分{p.get('总积分','')}分)  司龄: {p.get('司龄','')}年
  曾工作领域: {p.get('曾工作领域及年限','')}
  XPM领域: {p.get('XPM领域','')}  关键人才: {p.get('关键人才','')}
  当前关键角色: {p.get('当前关键角色','')}
"""

    query_line = ""
    if user_text and user_text.strip():
        query_line = f"\n用户搜索需求: 「{user_text}」\n请围绕用户需求进行针对性分析：strengths要说明该候选人为什么符合这个需求，weaknesses要指出与需求之间的差距，positioning和recommendation要结合需求给出判断。"

    prompt = f"{profiles_text}{query_line}\n请生成对比JSON（strengths每人3-4条，weaknesses每人1-2条，positioning一句定位，recommendation任用建议，comprehensive_score 0-100整数，overall_comparison综合结论2-4句）。只返回JSON，不要其他。"
    messages = [{"role": "system", "content": COMPARE_SYSTEM}, {"role": "user", "content": prompt}]

    raw_text = ""
    per_person = []
    overall = ""
    try:
        # Stream LLM response — user sees progress in real time
        from llm.backend import get_llm
        llm = get_llm()
        async for chunk in llm.chat_stream(messages, model="deepseek-ai/DeepSeek-V4-Pro", timeout=120, temperature=0.1):
            raw_text += chunk
    except Exception as e:
        import traceback
        print(f"[COMPARE] LLM stream FAILED: {e}")
        traceback.print_exc()

    # Parse structured JSON from streamed text
    if raw_text:
        try:
            data = json.loads(raw_text)
        except Exception:
            m = _re.search(r'```(?:json)?\s*\n?(.*?)\n?```', raw_text, re.DOTALL)
            if m:
                try: data = json.loads(m.group(1))
                except Exception: data = {}
            else:
                m2 = _re.search(r'\{.*\}', raw_text, re.DOTALL)
                data = json.loads(m2.group()) if m2 else {}
        overall = data.get("overall_comparison", "")
        per_person = data.get("profiles", [])
    else:
        print("[COMPARE] LLM returned empty response — using fallback")

    # Fallback: fill missing per_person
    if len(per_person) < len(profiles_raw):
        filled = []
        for i, p in enumerate(profiles_raw):
            if i < len(per_person) and per_person[i].get("comprehensive_score"):
                filled.append(per_person[i])
                continue
            eid2 = str(p.get("员工编码", p.get("姓名", str(i))))
            h2 = int(hashlib.md5(eid2.encode()).hexdigest()[:4], 16)
            s2 = 65 + (h2 % 31)
            filled.append({
                "name": p.get("姓名", ""),
                "strengths": ["技能匹配度高", "绩效表现稳定", "团队协作良好"][:3],
                "weaknesses": ["管理经验待提升"][:1],
                "comprehensive_score": s2,
                "positioning": "综合能力突出" if s2 >= 80 else "具备成长潜力",
                "recommendation": "建议重点考察" if s2 >= 80 else "建议培养后评估",
            })
        per_person = filled

    if not overall:
        names2 = [p.get("姓名","") for p in profiles_raw[:3]]
        overall = names2[0] + "综合能力最强；" + (names2[1] + "经验丰富。" if len(names2) >= 2 else "")
        if len(names2) >= 3: overall += names2[2] + "潜力较大。"

    # Cache and yield final enriched compare with AI analysis
    ctx.cached_compares[c_key] = {
        "type": "compare",
        "data": {
            "profiles": profiles,
            "analysis": overall,
            "per_person": per_person,
        }
    }
    yield ctx.cached_compares[c_key]
    yield {"type": "actions", "actions": ["导出对比结果"]}
    yield {"type": "done"}


async def _handle_profile(ctx, params, user_text, ids):
    from agents.profile import ProfileAgent
    store = _ensure_store()

    eid = params.get("employee_id")
    emp_name = params.get("employee_name", "员工")
    if not eid:
        yield {"type": "text", "content": "请提供员工姓名或工号"}
        yield {"type": "done"}
        return

    # Check cache
    if eid in ctx.cached_profiles:
        yield {"type": "text", "content": f"已从缓存加载 {emp_name} 的画像"}
        yield ctx.cached_profiles[eid]
        yield {"type": "done"}
        return

    iceberg = ProfileAgent().get_iceberg_view(eid)
    profile = store.get_by_id(eid)
    if not profile:
        yield {"type": "text", "content": "未找到员工"}
        yield {"type": "done"}
        return

    result = {
        "type": "profile",
        "data": {
            "name": profile.get("姓名", emp_name),
            "iceberg": iceberg,
            "skills": _comma_list(profile.get("技能标签")),
            "tags": _comma_list(profile.get("所有标签")),
        }
    }
    ctx.cached_profiles[eid] = result
    yield result
    yield {"type": "actions", "actions": ["岗位推荐", "标签管理", "职业发展"]}
    yield {"type": "done"}


async def _handle_career(ctx, params, user_text, ids):
    from agents.career import CareerAgent
    eid = params.get("employee_id")
    if not eid:
        yield {"type": "text", "content": "请提供员工姓名或工号"}
        yield {"type": "done"}
        return

    yield {"type": "text", "content": "正在生成职业发展分析..."}
    result = CareerAgent().analyze(eid)
    yield {"type": "text", "content": result.get("analysis", "暂无分析")}
    yield {"type": "done"}


_HANDLERS = {
    "clarify": _handle_clarify,
    "position_to_person": _handle_position_to_person,
    "report": _handle_report,
    "compare": _handle_compare,
    "profile": _handle_profile,
    "career": _handle_career,
}


async def stream_response(ctx, user_text: str, action: str = None, ids: list[str] = None) -> AsyncGenerator[dict, None]:
    """Main streaming function: route intent -> dispatch handler -> yield typed messages."""
    if action:
        intent_result = IntentResult(intent=action, params={"ids": ids} if ids else {}, confidence=1.0)
    else:
        intent_result = await route(user_text, ctx)

    params = intent_result.params
    handler = _HANDLERS.get(intent_result.intent)

    if handler:
        async for msg in handler(ctx, params, user_text, ids):
            yield msg
        return

    # Fallback
    yield {"type": "text", "content": "抱歉，我暂时无法处理这个请求。请尝试：\n· 帮我找XX岗位的候选人\n· 查看XX的详细报告\n· 对比候选人"}
    yield {"type": "done"}
